"""EHS ("Exploratory Hole Schedule") .xlsx parse stage.

Reads a landed EHS workbook into an in-memory, row-dict structure — mirrors
the AGS4/CPT-JSON parse stage's scope (tasks/plan/phase-3b-pipeline-
implementation.md Task 6): no validation, no transform, no DB writes. Raises
a clear, contextful error only for structurally unreadable input (missing
file, missing sheet, unexpected table header row) — never for bad *data*
values (a blank/malformed cell is returned as-is; catching that is the
semantic-validation stage's job, a later increment).

Grounded against the confirmed real layout in
geodb_etl.mappings.xlsx.ehs (HEADER_FIELD_CELLS/TABLE_COLUMNS/
TABLE_DATA_FIRST_ROW) — this module is the one place those constants are
actually read from a workbook; keep them in sync, don't duplicate the cell
map here.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from ...mappings.xlsx.ehs import (
    EHS_SHEET_NAME,
    HEADER_FIELD_CELLS,
    HEADER_VALUE_COLUMN,
    TABLE_COLUMNS,
    TABLE_DATA_FIRST_ROW,
    TABLE_HEADER_ROW,
)


class EhsParseError(Exception):
    """Raised when an EHS workbook cannot be structurally parsed.

    Always carries enough context (file path, and what was expected vs.
    found) to diagnose without re-opening the workbook by hand — per this
    repo's "include useful error messages with relevant context" standard.
    """


@dataclass(frozen=True)
class EhsDocument:
    """The parsed, unvalidated contents of one EHS workbook.

    Attributes:
        source_path: The workbook this document was parsed from.
        header: Raw header-block values keyed by label (see
            geodb_etl.mappings.xlsx.ehs.HEADER_FIELD_CELLS for the full set
            of labels) — values are whatever openpyxl returned for that
            cell (str, int, float, datetime, or None), not yet
            type-checked or required-ness-checked.
        hole_rows: One dict per tabular hole row, keyed by the table column
            labels in geodb_etl.mappings.xlsx.ehs.TABLE_COLUMNS, plus a
            "_row_number" key (the workbook's own 1-based row number) for
            RejectedRow traceability in the later validation stage.
    """

    source_path: Path
    header: dict[str, object]
    hole_rows: list[dict[str, object]] = field(default_factory=list)


def _normalize_header_text(value: object) -> str:
    """Collapse whitespace (including embedded newlines) for header comparison.

    The real workbook's table header row contains literal newlines inside
    cells (e.g. "Hole\\nnumber") — comparison against
    geodb_etl.mappings.xlsx.ehs.TABLE_COLUMNS's plain-text labels needs both
    sides normalized the same way.
    """

    return " ".join(str(value).split())


def parse_ehs_file(path: Path) -> EhsDocument:
    """Parse an EHS workbook into an EhsDocument.

    Args:
        path: Path to the .xlsx workbook (already landed — see a later "land"
            increment; this function does no hashing/archival itself).

    Returns:
        The parsed header block and hole rows, entirely unvalidated.

    Raises:
        EhsParseError: the file doesn't exist, isn't a readable .xlsx, is
            missing the expected "EHS" sheet, or that sheet's table header
            row (row 22) doesn't match the expected column labels — i.e. the
            file is structurally not an EHS workbook this pipeline
            recognises. Never raised for a blank/malformed data cell value.
    """

    if not path.is_file():
        raise EhsParseError(f"EHS workbook not found: {path}")

    try:
        workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except (InvalidFileException, BadZipFile, KeyError, OSError) as exc:
        raise EhsParseError(f"{path} is not a readable .xlsx workbook: {exc}") from exc

    try:
        if EHS_SHEET_NAME not in workbook.sheetnames:
            raise EhsParseError(
                f"{path} has no '{EHS_SHEET_NAME}' sheet "
                f"(found: {workbook.sheetnames})"
            )
        sheet = workbook[EHS_SHEET_NAME]

        header: dict[str, object] = {
            label: sheet[f"{HEADER_VALUE_COLUMN}{row}"].value
            for label, row in HEADER_FIELD_CELLS.items()
        }

        _check_table_header(sheet, path)
        hole_rows = _read_hole_rows(sheet)

        return EhsDocument(source_path=path, header=header, hole_rows=hole_rows)
    finally:
        workbook.close()


def _check_table_header(sheet: Any, path: Path) -> None:
    """Raise EhsParseError if the table header row doesn't match expectations.

    A structural check (this template's own shape), not a business-rule
    check — deliberately strict, since a mismatch here means either the
    wrong sheet/row was targeted or the EHS template itself changed in a way
    TABLE_COLUMNS hasn't been updated to reflect.
    """

    for label, column in TABLE_COLUMNS.items():
        cell_value = sheet[f"{column}{TABLE_HEADER_ROW}"].value
        if _normalize_header_text(cell_value) != label:
            raise EhsParseError(
                f"{path}: expected table header '{label}' at "
                f"{column}{TABLE_HEADER_ROW}, found {cell_value!r}"
            )


def _read_hole_rows(sheet: Any) -> list[dict[str, object]]:
    """Read tabular hole rows starting at TABLE_DATA_FIRST_ROW.

    Terminates at the first row whose "No." column is blank — the schedule's
    length varies per campaign (not a fixed row count), and the real
    reference workbook confirms a blank row (then an unrelated "Summary"
    block) immediately follows the last hole row.
    """

    no_column = TABLE_COLUMNS["No."]
    rows: list[dict[str, object]] = []
    row_number = TABLE_DATA_FIRST_ROW
    while sheet[f"{no_column}{row_number}"].value not in (None, ""):
        row_data: dict[str, object] = {
            label: sheet[f"{column}{row_number}"].value
            for label, column in TABLE_COLUMNS.items()
        }
        row_data["_row_number"] = row_number
        rows.append(row_data)
        row_number += 1

    return rows

